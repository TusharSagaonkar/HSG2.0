import re
from datetime import datetime
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from .base import BaseStatementParser, ParseResult, ParsedTransaction


class XLSXParser(BaseStatementParser):
    """
    Parser for XLSX (Excel) bank statements.

    Supports:
      - HDFC Bank XLSX format
      - ICICI Bank XLSX format
      - SBI XLSX format
      - Generic Excel exports

    Uses openpyxl for reading .xlsx files.
    """

    bank_name = "Generic XLSX"
    _supported_extensions = ("xlsx", "xls")

    # Known header patterns for specific bank formats
    BANK_FORMAT_PATTERNS = {
        "HDFC": {
            "date": ["Date", "Txn Date", "Transaction Date"],
            "narration": ["Narration", "Description", "Particulars"],
            "cheque_no": ["Chq./Ref.No.", "Cheque No", "Cheque Number"],
            "debit": ["Withdrawal Amt.", "Debit", "Withdrawal"],
            "credit": ["Deposit Amt.", "Credit", "Deposit"],
            "balance": ["Closing Balance", "Balance"],
        },
        "ICICI": {
            "date": ["Date", "Transaction Date", "Value Date"],
            "narration": ["Description", "Narration", "Particulars"],
            "reference_no": ["Ref No.", "Transaction Reference", "Reference No"],
            "debit": ["Debit", "Withdrawal"],
            "credit": ["Credit", "Deposit"],
            "balance": ["Balance", "Available Balance"],
        },
        "SBI": {
            "date": ["Txn Date", "Date", "Transaction Date"],
            "narration": ["Description", "Narration", "Particulars"],
            "reference_no": ["Ref No.", "Ref.No.", "Reference"],
            "debit": ["Debit", "Withdrawal"],
            "credit": ["Credit", "Deposit"],
            "balance": ["Balance", "Available Balance"],
        },
    }

    DATE_FORMATS = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d %b %Y",
        "%d %B %Y",
    ]

    SUMMARY_KEYWORDS = {"total", "closing", "opening", "b/f", "c/f", "balance c/f",
                        "balance b/f", "grand total", "opening balance", "closing balance"}

    def parse(self, file_obj, filename: str = "") -> ParseResult:
        result = ParseResult(transactions=[], parser_name="XLSX")

        if openpyxl is None:
            result.errors.append(
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )
            return result

        try:
            workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception as e:
            result.errors.append(f"Failed to open XLSX file: {e}")
            return result

        all_transactions = []
        sheet_count = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(min_row=1, values_only=True))
            if not rows:
                continue

            sheet_count += 1
            header_idx, column_map = self._find_header(rows)

            if header_idx is None:
                result.warnings.append(
                    f"Sheet '{sheet_name}': Could not detect header row. Skipping."
                )
                continue

            for row_idx in range(header_idx + 1, len(rows)):
                row = rows[row_idx]
                if self._is_empty_row(row):
                    continue

                if self._is_summary_row(row):
                    continue

                txn = self._parse_row(row, column_map)
                if txn:
                    all_transactions.append(txn)
                else:
                    result.warnings.append(
                        f"Sheet '{sheet_name}' row {row_idx + 1}: Skipped — missing date or amount."
                    )

        workbook.close()

        if sheet_count == 0:
            result.errors.append("No sheets found in the workbook.")
            return result

        result.transactions = all_transactions

        if result.transactions:
            dates = sorted(t.transaction_date for t in result.transactions)
            result.statement_start_date = dates[0]
            result.statement_end_date = dates[-1]

        if not result.transactions and not result.errors:
            result.errors.append("No valid transactions found in any sheet.")

        return result

    def _is_summary_row(self, row: tuple) -> bool:
        """Detect summary/footer rows like TOTAL, CLOSING, etc."""
        if not row:
            return False
        first_cell = str(row[0]).strip().lower() if row[0] is not None else ""
        return first_cell in self.SUMMARY_KEYWORDS

    def _find_header(self, rows: list[tuple]) -> tuple[Optional[int], dict]:
        # Try known bank formats first
        for bank_name, patterns in self.BANK_FORMAT_PATTERNS.items():
            for idx in range(min(15, len(rows))):
                row = rows[idx]
                col_map = self._match_header_row(row, patterns)
                if len(col_map) >= 3 and "date" in col_map:
                    return idx, col_map

        # Generic fallback: look for common keywords
        for idx in range(min(15, len(rows))):
            row = rows[idx]
            col_map = {}
            for col_idx, cell in enumerate(row):
                cell_str = self._clean_text(str(cell) if cell is not None else "")
                if not cell_str:
                    continue
                cell_lower = cell_str.lower()

                if any(p in cell_lower for p in ("date",)):
                    col_map.setdefault("date", col_idx)
                elif any(p in cell_lower for p in ("narration", "description", "particulars", "remarks")):
                    col_map.setdefault("narration", col_idx)
                elif any(p in cell_lower for p in ("ref", "utr", "reference")):
                    col_map.setdefault("reference_no", col_idx)
                elif any(p in cell_lower for p in ("cheque", "cheque", "chq")):
                    col_map.setdefault("cheque_no", col_idx)
                elif any(p in cell_lower for p in ("debit", "withdrawal", "dr")):
                    col_map.setdefault("debit", col_idx)
                elif any(p in cell_lower for p in ("credit", "deposit", "cr")):
                    col_map.setdefault("credit", col_idx)
                elif any(p in cell_lower for p in ("balance",)):
                    col_map.setdefault("balance", col_idx)

            if len(col_map) >= 3 and "date" in col_map:
                return idx, col_map

        return None, {}

    def _match_header_row(self, row: tuple, patterns: dict) -> dict:
        col_map = {}
        for col_idx, cell in enumerate(row):
            cell_str = self._clean_text(str(cell) if cell is not None else "")
            if not cell_str:
                continue

            for col_name, pattern_list in patterns.items():
                for pattern in pattern_list:
                    if cell_str.lower() == pattern.lower():
                        col_map.setdefault(col_name, col_idx)
                        break

        return col_map

    def _parse_row(
        self,
        row: tuple,
        column_map: dict,
    ) -> Optional[ParsedTransaction]:
        cells = list(row)

        date_raw = self._get_cell(cells, column_map, "date")
        if not date_raw:
            return None

        date_val = self._parse_date(date_raw)
        if not date_val:
            return None

        narration_val = self._get_cell(cells, column_map, "narration")
        if not narration_val:
            narration_val = self._get_cell(cells, column_map, "reference_no")

        debit_val = self._get_cell(cells, column_map, "debit")
        credit_val = self._get_cell(cells, column_map, "credit")

        amount, dr_cr = self._detect_dr_cr(debit_val, credit_val)

        if amount <= 0:
            return None

        return ParsedTransaction(
            transaction_date=date_val,
            narration=self._clean_text(narration_val),
            amount=amount,
            dr_cr=dr_cr,
            reference_no=self._clean_text(
                self._get_cell(cells, column_map, "reference_no")
            ),
            cheque_no=self._clean_text(
                self._get_cell(cells, column_map, "cheque_no")
            ),
            value_date=None,
            balance=self._clean_text(
                self._get_cell(cells, column_map, "balance")
            ) or None,
            raw_row={
                str(idx): self._clean_text(str(c) if c is not None else "")
                for idx, c in enumerate(cells)
            },
        )

    def _parse_date(self, value) -> Optional[str]:
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        value_str = str(value).strip()

        for fmt in self.DATE_FORMATS:
            try:
                parsed = datetime.strptime(value_str, fmt)
                return parsed.strftime("%Y-%m-%d")
            except ValueError:
                continue

        for sep in ("-", "/", " "):
            if sep in value_str:
                try:
                    parts = value_str.split(sep)
                    if len(parts) == 3:
                        parsed = datetime(
                            int(parts[0]),
                            int(parts[1]),
                            int(parts[2]),
                        )
                        return parsed.strftime("%Y-%m-%d")
                except (ValueError, IndexError):
                    pass

        return None

    @staticmethod
    def _get_cell(cells: list, column_map: dict, key: str) -> str:
        idx = column_map.get(key)
        if idx is not None and idx < len(cells):
            val = cells[idx]
            return str(val) if val is not None else ""
        return ""

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(
            cell is None or (isinstance(cell, str) and not cell.strip())
            for cell in row
        )