from dataclasses import dataclass
import csv
import io
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore


@dataclass(frozen=True)
class DetectionResult:
    bank_name: str
    format_name: str
    file_type: str
    confidence: int


class StatementFormatDetector:
    """Lightweight detector for file type and bank format."""

    def detect(self, filename: str, header_row=None, content: bytes | None = None) -> DetectionResult:
        file_type = self._detect_file_type(filename)
        bank_name = "Generic"
        format_name = f"GENERIC_{file_type.upper()}"
        confidence = 50

        header_text = self._normalize_header_text(header_row, content, file_type)
        for candidate in ("hdfc", "icici", "sbi", "axis", "kotak", "idfc"):
            if candidate in header_text:
                bank_name = candidate.upper()
                format_name = f"{bank_name}_{file_type.upper()}"
                confidence = 85
                break

        return DetectionResult(
            bank_name=bank_name,
            format_name=format_name,
            file_type=file_type,
            confidence=confidence,
        )

    @staticmethod
    def _detect_file_type(filename: str) -> str:
        suffix = Path(filename).suffix.lower().lstrip(".")
        if suffix in {"csv", "txt", "xls", "xlsx", "pdf", "json", "xml"}:
            return suffix
        return "unknown"

    def _normalize_header_text(self, header_row, content: bytes | None, file_type: str) -> str:
        parts = [str(cell).lower() for cell in (header_row or []) if cell is not None]

        if content:
            if file_type in {"csv", "txt"}:
                try:
                    text = content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    text = content.decode("latin-1", errors="ignore")
                first_line = text.splitlines()[0] if text.splitlines() else ""
                parts.append(first_line.lower())
                try:
                    reader = csv.reader(io.StringIO(text))
                    rows = list(reader)[:3]
                    for row in rows:
                        parts.extend(str(cell).lower() for cell in row if cell is not None)
                except Exception:
                    pass
            elif file_type in {"xlsx", "xls"} and openpyxl is not None:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    for sheet_name in wb.sheetnames[:2]:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                            parts.extend(str(cell).lower() for cell in row if cell is not None)
                    wb.close()
                except Exception:
                    pass

        return " ".join(parts)
