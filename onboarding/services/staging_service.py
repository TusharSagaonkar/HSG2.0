"""Staging area service for the Society Creation & Accounting Migration Wizard
(Phase 4 — Steps 10–14: file upload, parsing, storage, deletion, approval).

This service is the **single authority** over the staging tables. It manages
the lifecycle of uploaded migration data:

1. ``upload_file`` — save the uploaded file, create an :class:`UploadBatch`,
   parse rows, and persist them into the appropriate staging model.
2. ``parse_file`` — detect file type (CSV/XLSX) and return a list of row dicts.
3. ``store_staging_data`` — map parsed rows to the correct staging model.
4. ``get_staging_data`` — retrieve staged rows + validation status counts.
5. ``delete_batch`` — remove staged data (supports delete & re-upload, R-5).
6. ``approve_batch`` — lock staged data as approved (only if all rows VALID).
7. ``get_upload_summary`` — per-template upload/validation/approval summary.

Design notes
------------
- **Never writes to live accounting tables.** This service only touches the
  ``onboarding`` staging models and :class:`UploadBatch`.
- **All mutations are ``@transaction.atomic``** per the service contract.
- **Tenant safety:** staging models use :class:`TenantManager` with a
  ``society`` FK (SET_NULL, nullable). To avoid surprises when the tenant
  contextvar is unset or set to a different society, all reads use
  ``.unscoped().filter(wizard=wizard, ...)`` and filter explicitly by wizard.
- **Re-upload support:** ``upload_file`` deletes any previous staging data
  for the same template_type before storing the new upload (R-5).
- **Audit robustness:** audit-log writes are wrapped so a logging failure
  never blocks a legitimate staging operation (pattern from
  :mod:`onboarding.services.wizard_service`).
- **Excel support:** uses ``openpyxl`` if installed; otherwise only CSV is
  supported and ``.xlsx`` uploads raise a clear ``ValueError``.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.files.storage import default_storage
from django.db import transaction

from onboarding.models import (
    MigrationAuditLog,
    OnboardingWizard,
    StagingBankOpening,
    StagingCashOpening,
    StagingChartOfAccounts,
    StagingFixedAsset,
    StagingFund,
    StagingLoan,
    StagingMemberOutstanding,
    StagingSecurityDeposit,
    StagingTrialBalance,
    StagingVendorOutstanding,
    UploadBatch,
)

# openpyxl is optional. If absent, only CSV uploads are supported.
try:  # pragma: no cover - import guard
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------- #
# Decimal tolerance / defaults
# --------------------------------------------------------------------------- #
ZERO = Decimal("0")
TOLERANCE = Decimal("0.01")

# --------------------------------------------------------------------------- #
# Template type → staging model mapping
# --------------------------------------------------------------------------- #
# Canonical template keys (match UploadBatch.TemplateType values).
TEMPLATE_CHART_OF_ACCOUNTS = "CHART_OF_ACCOUNTS"
TEMPLATE_TRIAL_BALANCE = "TRIAL_BALANCE"
TEMPLATE_MEMBER_OUTSTANDING = "MEMBER_OUTSTANDING"
TEMPLATE_VENDOR_OUTSTANDING = "VENDOR_OUTSTANDING"
TEMPLATE_BANK_OPENING = "BANK_OPENING"
TEMPLATE_CASH_OPENING = "CASH_OPENING"
TEMPLATE_FIXED_ASSETS = "FIXED_ASSETS"
TEMPLATE_SECURITY_DEPOSITS = "SECURITY_DEPOSITS"
TEMPLATE_LOANS = "LOANS"
TEMPLATE_FUNDS = "FUNDS"

# Ordered list of all canonical template types.
ALL_TEMPLATE_TYPES = (
    TEMPLATE_CHART_OF_ACCOUNTS,
    TEMPLATE_TRIAL_BALANCE,
    TEMPLATE_MEMBER_OUTSTANDING,
    TEMPLATE_VENDOR_OUTSTANDING,
    TEMPLATE_BANK_OPENING,
    TEMPLATE_CASH_OPENING,
    TEMPLATE_FIXED_ASSETS,
    TEMPLATE_SECURITY_DEPOSITS,
    TEMPLATE_LOANS,
    TEMPLATE_FUNDS,
)

# Maps canonical template type → staging model class.
TEMPLATE_MODEL_MAP: dict[str, type] = {
    TEMPLATE_CHART_OF_ACCOUNTS: StagingChartOfAccounts,
    TEMPLATE_TRIAL_BALANCE: StagingTrialBalance,
    TEMPLATE_MEMBER_OUTSTANDING: StagingMemberOutstanding,
    TEMPLATE_VENDOR_OUTSTANDING: StagingVendorOutstanding,
    TEMPLATE_BANK_OPENING: StagingBankOpening,
    TEMPLATE_CASH_OPENING: StagingCashOpening,
    TEMPLATE_FIXED_ASSETS: StagingFixedAsset,
    TEMPLATE_SECURITY_DEPOSITS: StagingSecurityDeposit,
    TEMPLATE_LOANS: StagingLoan,
    TEMPLATE_FUNDS: StagingFund,
}

# Aliases accepted as template_type arguments (T1, T1_CHART_OF_ACCOUNTS, etc.).
# Normalized to the canonical key.
_TEMPLATE_ALIASES: dict[str, str] = {
    "T1": TEMPLATE_CHART_OF_ACCOUNTS,
    "T1_CHART_OF_ACCOUNTS": TEMPLATE_CHART_OF_ACCOUNTS,
    "CHART_OF_ACCOUNTS": TEMPLATE_CHART_OF_ACCOUNTS,
    "T2": TEMPLATE_TRIAL_BALANCE,
    "T2_TRIAL_BALANCE": TEMPLATE_TRIAL_BALANCE,
    "TRIAL_BALANCE": TEMPLATE_TRIAL_BALANCE,
    "T3": TEMPLATE_MEMBER_OUTSTANDING,
    "T3_MEMBER_OUTSTANDING": TEMPLATE_MEMBER_OUTSTANDING,
    "MEMBER_OUTSTANDING": TEMPLATE_MEMBER_OUTSTANDING,
    "T4": TEMPLATE_VENDOR_OUTSTANDING,
    "T4_VENDOR_OUTSTANDING": TEMPLATE_VENDOR_OUTSTANDING,
    "VENDOR_OUTSTANDING": TEMPLATE_VENDOR_OUTSTANDING,
    "T5": TEMPLATE_BANK_OPENING,
    "T5_BANK_OPENING": TEMPLATE_BANK_OPENING,
    "BANK_OPENING": TEMPLATE_BANK_OPENING,
    "T6": TEMPLATE_CASH_OPENING,
    "T6_CASH_OPENING": TEMPLATE_CASH_OPENING,
    "CASH_OPENING": TEMPLATE_CASH_OPENING,
    "T7": TEMPLATE_FIXED_ASSETS,
    "T7_FIXED_ASSETS": TEMPLATE_FIXED_ASSETS,
    "FIXED_ASSETS": TEMPLATE_FIXED_ASSETS,
    "T8": TEMPLATE_SECURITY_DEPOSITS,
    "T8_SECURITY_DEPOSITS": TEMPLATE_SECURITY_DEPOSITS,
    "SECURITY_DEPOSITS": TEMPLATE_SECURITY_DEPOSITS,
    "T9": TEMPLATE_LOANS,
    "T9_LOANS": TEMPLATE_LOANS,
    "LOANS": TEMPLATE_LOANS,
    "T10": TEMPLATE_FUNDS,
    "T10_FUNDS": TEMPLATE_FUNDS,
    "FUNDS": TEMPLATE_FUNDS,
}

# --------------------------------------------------------------------------- #
# Per-template column definitions (expected column names, snake_case)
# --------------------------------------------------------------------------- #
TEMPLATE_COLUMNS: dict[str, tuple[str, ...]] = {
    TEMPLATE_CHART_OF_ACCOUNTS: (
        "account_code", "account_name", "account_group", "account_type",
        "parent_code", "nature", "opening_debit", "opening_credit",
    ),
    TEMPLATE_TRIAL_BALANCE: ("account_code", "account_name", "debit", "credit"),
    TEMPLATE_MEMBER_OUTSTANDING: (
        "unit_identifier", "member_name", "outstanding_amount",
        "advance_maintenance", "credit_balance", "late_fees",
        "interest_receivable",
    ),
    TEMPLATE_VENDOR_OUTSTANDING: (
        "vendor_name", "outstanding_amount", "advance_paid", "retention",
        "security_deposit",
    ),
    TEMPLATE_BANK_OPENING: (
        "bank_name", "account_number", "ifsc", "branch", "opening_balance",
        "account_code",
    ),
    TEMPLATE_CASH_OPENING: ("opening_balance",),
    TEMPLATE_FIXED_ASSETS: (
        "asset_name", "asset_category", "gross_value", "depreciation",
        "net_value", "account_code",
    ),
    TEMPLATE_SECURITY_DEPOSITS: ("description", "amount", "against_account"),
    TEMPLATE_LOANS: (
        "loan_name", "loan_type", "outstanding_principal", "interest",
        "account_code",
    ),
    TEMPLATE_FUNDS: ("fund_name", "fund_type", "balance", "account_code"),
}

# Per-template decimal field names (used for safe parsing during storage).
TEMPLATE_DECIMAL_FIELDS: dict[str, tuple[str, ...]] = {
    TEMPLATE_CHART_OF_ACCOUNTS: ("opening_debit", "opening_credit"),
    TEMPLATE_TRIAL_BALANCE: ("debit", "credit"),
    TEMPLATE_MEMBER_OUTSTANDING: (
        "outstanding_amount", "advance_maintenance", "credit_balance",
        "late_fees", "interest_receivable",
    ),
    TEMPLATE_VENDOR_OUTSTANDING: (
        "outstanding_amount", "advance_paid", "retention", "security_deposit",
    ),
    TEMPLATE_BANK_OPENING: ("opening_balance",),
    TEMPLATE_CASH_OPENING: ("opening_balance",),
    TEMPLATE_FIXED_ASSETS: ("gross_value", "depreciation", "net_value"),
    TEMPLATE_SECURITY_DEPOSITS: ("amount",),
    TEMPLATE_LOANS: ("outstanding_principal", "interest"),
    TEMPLATE_FUNDS: ("balance",),
}

# Fields excluded from row serialization (FKs / internal).
_SERIALIZE_EXCLUDE = {"wizard", "society", "upload_batch", "id"}


class StagingService:
    """Manages the migration staging area: upload, parse, store, delete,
    approve. Never writes to live accounting tables.
    """

    # ------------------------------------------------------------------ #
    # Template helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _get_staging_model(template_type):
        """Map a template_type string to the staging model class.

        Accepts any alias (``T1``, ``T1_CHART_OF_ACCOUNTS``,
        ``CHART_OF_ACCOUNTS``). Raises ``ValueError`` for unknown types.
        """
        canonical = StagingService._normalize_template_type(template_type)
        return TEMPLATE_MODEL_MAP[canonical]

    @staticmethod
    def _normalize_template_type(template_type) -> str:
        """Normalize any accepted template_type alias to its canonical key."""
        if template_type is None:
            raise ValueError("template_type is required.")
        key = str(template_type).strip().upper()
        canonical = _TEMPLATE_ALIASES.get(key)
        if canonical is None:
            raise ValueError(
                f"Unknown template_type '{template_type}'. "
                f"Expected one of: {sorted(set(_TEMPLATE_ALIASES))}"
            )
        return canonical

    @staticmethod
    def get_template_columns(template_type) -> tuple[str, ...]:
        """Return the expected column names (snake_case) for a template type."""
        canonical = StagingService._normalize_template_type(template_type)
        return TEMPLATE_COLUMNS[canonical]

    # ------------------------------------------------------------------ #
    # File parsing
    # ------------------------------------------------------------------ #

    @staticmethod
    def parse_file(file_path) -> list[dict]:
        """Parse a CSV or XLSX file into a list of row dicts.

        Detects the file type by extension. For CSV uses
        :class:`csv.DictReader`; for XLSX uses ``openpyxl`` (if installed).

        The returned dicts use the **original header strings** as keys (not
        normalized) so callers can preserve the raw row. Header normalization
        happens during storage.

        Raises ``ValueError`` if the file cannot be read, is empty, or is an
        unsupported type.
        """
        if not file_path:
            raise ValueError("file_path is required.")

        # Resolve to an absolute filesystem path (handles media-relative paths).
        abs_path = file_path
        if not os.path.exists(abs_path):
            try:
                abs_path = default_storage.path(file_path)
            except Exception:  # noqa: BLE001 — fall through to existence check
                abs_path = file_path
        if not os.path.exists(abs_path):
            raise ValueError(f"File not found: {file_path}")

        ext = os.path.splitext(abs_path)[1].lower()
        if ext == ".csv":
            return StagingService._parse_csv(abs_path)
        if ext in (".xlsx", ".xlsm"):
            return StagingService._parse_xlsx(abs_path)
        if ext == ".xls":
            raise ValueError(
                "Legacy .xls format is not supported. Please save the file as .xlsx or .csv."
            )
        raise ValueError(
            f"Unsupported file extension '{ext}'. Only .csv and .xlsx are supported."
        )

    @staticmethod
    def _parse_csv(abs_path: str) -> list[dict]:
        rows: list[dict] = []
        # Try common encodings for Indian CSV exports.
        text = StagingService._read_text(abs_path)
        if not text.strip():
            raise ValueError("CSV file is empty.")

        # Auto-detect delimiter (comma vs tab vs semicolon).
        delimiter = StagingService._detect_csv_delimiter(text)

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValueError("CSV file has no header row.")

        for row in reader:
            # Skip fully-empty rows.
            if not any((v or "").strip() for v in row.values()):
                continue
            rows.append({k: ("" if v is None else v) for k, v in row.items()})
        return rows

    @staticmethod
    def _parse_xlsx(abs_path: str) -> list[dict]:
        if openpyxl is None:  # pragma: no cover
            raise ValueError(
                "Excel (.xlsx) parsing requires the 'openpyxl' package, which is "
                "not installed. Install it with: pip install openpyxl — or upload "
                "a .csv file instead."
            )
        rows: list[dict] = []
        workbook = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            all_rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not all_rows:
            raise ValueError("XLSX file is empty.")

        headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        if not any(headers):
            raise ValueError("XLSX file has no header row.")

        for raw in all_rows[1:]:
            if raw is None:
                continue
            # Skip fully-empty rows.
            if not any((c is not None and str(c).strip()) for c in raw):
                continue
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = raw[idx] if idx < len(raw) else None
                row[header] = "" if value is None else str(value)
            rows.append(row)
        return rows

    @staticmethod
    def _read_text(abs_path: str) -> str:
        raw = open(abs_path, "rb").read()
        for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")

    @staticmethod
    def _detect_csv_delimiter(text: str) -> str:
        first_line = text.splitlines()[0] if text else ""
        if "\t" in first_line:
            return "\t"
        if ";" in first_line and "," not in first_line:
            return ";"
        return ","

    # ------------------------------------------------------------------ #
    # Upload + store
    # ------------------------------------------------------------------ #

    @staticmethod
    @transaction.atomic
    def upload_file(wizard, template_type, file, user=None) -> UploadBatch:
        """Save an uploaded file, create an :class:`UploadBatch`, parse it,
        and store rows into the appropriate staging table.

        Supports re-upload: any previous staging data for the same
        template_type is deleted before storing the new data (R-5).

        Returns the created :class:`UploadBatch`.
        """
        if wizard is None:
            raise ValueError("wizard is required.")
        if file is None:
            raise ValueError("file is required.")

        canonical = StagingService._normalize_template_type(template_type)
        model_cls = TEMPLATE_MODEL_MAP[canonical]

        # Validate file extension.
        file_name = getattr(file, "name", "") or ""
        ext = os.path.splitext(file_name)[1].lower()
        if ext not in (".csv", ".xlsx", ".xlsm"):
            raise ValueError(
                f"Unsupported file extension '{ext}'. Only .csv and .xlsx are supported."
            )
        if ext in (".xlsx", ".xlsm") and openpyxl is None:  # pragma: no cover
            raise ValueError(
                "Excel (.xlsx) parsing requires the 'openpyxl' package, which is "
                "not installed. Please upload a .csv file instead."
            )

        # Persist the uploaded file to media storage.
        safe_name = os.path.basename(file_name) or f"upload{ext}"
        rel_dir = f"onboarding/staging/{wizard.pk}/{canonical}"
        saved_path = default_storage.save(
            os.path.join(rel_dir, safe_name), file
        )

        # Parse the saved file into row dicts.
        rows = StagingService.parse_file(saved_path)

        # Delete any previous staging data + batch for this template (re-upload).
        StagingService._delete_existing(wizard, canonical)

        # Create the UploadBatch record.
        batch = UploadBatch.objects.create(
            wizard=wizard,
            society=wizard.society,
            template_type=canonical,
            file_name=safe_name,
            file_path=saved_path,
            uploaded_by=user,
            row_count=len(rows),
            status=UploadBatch.Status.UPLOADED,
            validation_summary={
                "total_rows": len(rows),
                "valid_rows": 0,
                "invalid_rows": 0,
                "pending_rows": len(rows),
                "errors_count": 0,
            },
        )

        # Store rows into the staging table.
        StagingService.store_staging_data(wizard, batch, canonical, rows)

        StagingService._log_audit(
            wizard=wizard,
            action="UPLOAD",
            user=user,
            details={
                "template_type": canonical,
                "file_name": safe_name,
                "file_path": saved_path,
                "row_count": len(rows),
                "batch_id": batch.pk,
            },
        )
        return batch

    @staticmethod
    @transaction.atomic
    def store_staging_data(wizard, upload_batch, template_type, rows) -> int:
        """Map parsed rows to the appropriate staging model and bulk-create
        them. Each row gets ``row_number``, ``raw_data`` (original row dict),
        and ``validation_status=PENDING``.

        Returns the number of staging rows created.
        """
        canonical = StagingService._normalize_template_type(template_type)
        model_cls = TEMPLATE_MODEL_MAP[canonical]
        decimal_fields = TEMPLATE_DECIMAL_FIELDS[canonical]
        columns = TEMPLATE_COLUMNS[canonical]

        society = wizard.society
        objects_to_create = []
        for idx, raw_row in enumerate(rows, start=1):
            normalized = StagingService._normalize_row(raw_row)

            kwargs: dict[str, Any] = {
                "wizard": wizard,
                "society": society,
                "upload_batch": upload_batch,
                "row_number": idx,
                "raw_data": StagingService._json_safe_row(raw_row),
                "validation_status": model_cls.ValidationStatus.PENDING,
                "validation_errors": [],
                "is_approved": False,
            }

            # Populate model-specific fields from the normalized row.
            for col in columns:
                value = normalized.get(col, "")
                if col in decimal_fields:
                    kwargs[col] = StagingService._to_decimal(value)
                else:
                    kwargs[col] = str(value).strip() if value is not None else ""
            objects_to_create.append(model_cls(**kwargs))

        if objects_to_create:
            model_cls.objects.bulk_create(objects_to_create, batch_size=500)

        # Keep the batch row_count in sync (in case store is called directly).
        if upload_batch is not None and upload_batch.pk:
            upload_batch.row_count = len(objects_to_create)
            upload_batch.save(update_fields=["row_count"])

        return len(objects_to_create)

    # ------------------------------------------------------------------ #
    # Retrieval
    # ------------------------------------------------------------------ #

    @staticmethod
    def get_staging_data(wizard, template_type) -> dict:
        """Return all staging rows for the given template_type and wizard,
        ordered by row_number.

        Returns a dict::

            {
                "rows": [...],          # list of serialized row dicts
                "total_count": int,
                "valid_count": int,
                "invalid_count": int,
                "pending_count": int,
            }
        """
        canonical = StagingService._normalize_template_type(template_type)
        model_cls = TEMPLATE_MODEL_MAP[canonical]

        qs = (
            model_cls.objects
            .filter(wizard=wizard)
            .order_by("row_number")
        )
        rows = [StagingService._serialize_row(r) for r in qs]

        counts = StagingService._status_counts(qs)
        return {
            "rows": rows,
            "total_count": counts["total"],
            "valid_count": counts["valid"],
            "invalid_count": counts["invalid"],
            "pending_count": counts["pending"],
        }

    @staticmethod
    def get_upload_summary(wizard) -> dict:
        """Return a summary of all template types and their upload status.

        For each template_type: whether data exists, row count, validation
        status counts, and approval status (UploadBatch.status).
        """
        summary: dict[str, Any] = {}
        for canonical in ALL_TEMPLATE_TYPES:
            model_cls = TEMPLATE_MODEL_MAP[canonical]
            qs = model_cls.objects.filter(wizard=wizard)
            counts = StagingService._status_counts(qs)

            batch = (
                UploadBatch.objects.unscoped()
                .filter(wizard=wizard, template_type=canonical)
                .order_by("-uploaded_at")
                .first()
            )

            summary[canonical] = {
                "has_data": counts["total"] > 0,
                "row_count": counts["total"],
                "valid_count": counts["valid"],
                "invalid_count": counts["invalid"],
                "pending_count": counts["pending"],
                "approval_status": batch.status if batch else None,
                "batch_id": batch.pk if batch else None,
                "file_name": batch.file_name if batch else None,
                "uploaded_at": batch.uploaded_at.isoformat() if batch else None,
            }
        return summary

    # ------------------------------------------------------------------ #
    # Delete + approve
    # ------------------------------------------------------------------ #

    @staticmethod
    @transaction.atomic
    def delete_batch(wizard, template_type, user=None) -> None:
        """Delete all staging rows for the template_type and the associated
        :class:`UploadBatch`. Sets ``UploadBatch.status=DELETED`` before
        deleting. Logs via :class:`MigrationAuditLog`.

        Supports the "delete and re-upload" requirement (R-5).
        """
        canonical = StagingService._normalize_template_type(template_type)
        model_cls = TEMPLATE_MODEL_MAP[canonical]

        qs = model_cls.objects.filter(wizard=wizard)
        row_count = qs.count()

        batches = list(
            UploadBatch.objects.unscoped()
            .filter(wizard=wizard, template_type=canonical)
        )

        # Mark batches DELETED before deleting (for audit snapshot).
        for batch in batches:
            batch.status = UploadBatch.Status.DELETED
            batch.save(update_fields=["status"])

        StagingService._log_audit(
            wizard=wizard,
            action="DELETE",
            user=user,
            details={
                "template_type": canonical,
                "row_count": row_count,
                "batch_ids": [b.pk for b in batches],
            },
        )

        # Delete staging rows, then the batches.
        qs.delete()
        for batch in batches:
            batch.delete()

    @staticmethod
    @transaction.atomic
    def approve_batch(wizard, template_type, user=None) -> UploadBatch:
        """Mark all staging rows as ``is_approved=True`` and set the
        :class:`UploadBatch` status to ``APPROVED``.

        Only allowed if **all** rows have ``validation_status=VALID``.
        Raises ``ValueError`` if any rows are INVALID or PENDING.
        """
        canonical = StagingService._normalize_template_type(template_type)
        model_cls = TEMPLATE_MODEL_MAP[canonical]

        qs = model_cls.objects.filter(wizard=wizard)
        total = qs.count()
        if total == 0:
            raise ValueError(
                f"No staging data found for template '{canonical}' to approve."
            )

        invalid = qs.filter(
            validation_status=model_cls.ValidationStatus.INVALID
        ).count()
        pending = qs.filter(
            validation_status=model_cls.ValidationStatus.PENDING
        ).count()
        if invalid or pending:
            raise ValueError(
                f"Cannot approve template '{canonical}': "
                f"{invalid} invalid row(s) and {pending} pending row(s). "
                "All rows must be VALID before approval."
            )

        qs.update(is_approved=True)

        batch = (
            UploadBatch.objects.unscoped()
            .filter(wizard=wizard, template_type=canonical)
            .order_by("-uploaded_at")
            .first()
        )
        if batch is not None:
            batch.status = UploadBatch.Status.APPROVED
            batch.save(update_fields=["status"])

        StagingService._log_audit(
            wizard=wizard,
            action="APPROVE",
            user=user,
            details={
                "template_type": canonical,
                "row_count": total,
                "batch_id": batch.pk if batch else None,
            },
        )
        return batch

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _delete_existing(wizard, canonical: str) -> None:
        """Delete existing staging rows + batches for a template (re-upload)."""
        model_cls = TEMPLATE_MODEL_MAP[canonical]
        qs = model_cls.objects.filter(wizard=wizard)
        qs.delete()
        UploadBatch.objects.unscoped().filter(
            wizard=wizard, template_type=canonical
        ).delete()

    @staticmethod
    def _status_counts(qs) -> dict:
        """Return validation status counts for a staging queryset."""
        # Use the model's ValidationStatus choices generically.
        status_field = qs.model._meta.get_field("validation_status")
        valid_val = next(
            (v for v, _ in status_field.choices if v == "VALID"), "VALID"
        )
        invalid_val = next(
            (v for v, _ in status_field.choices if v == "INVALID"), "INVALID"
        )
        pending_val = next(
            (v for v, _ in status_field.choices if v == "PENDING"), "PENDING"
        )
        total = qs.count()
        valid = qs.filter(validation_status=valid_val).count()
        invalid = qs.filter(validation_status=invalid_val).count()
        pending = qs.filter(validation_status=pending_val).count()
        return {
            "total": total,
            "valid": valid,
            "invalid": invalid,
            "pending": pending,
        }

    @staticmethod
    def _serialize_row(instance) -> dict:
        """Serialize a staging row instance into a JSON-safe dict."""
        data: dict[str, Any] = {}
        for field in instance._meta.get_fields():
            name = getattr(field, "name", None)
            if name is None or name in _SERIALIZE_EXCLUDE:
                continue
            value = getattr(instance, name, None)
            if isinstance(value, Decimal):
                data[name] = str(value)
            elif hasattr(value, "isoformat"):
                data[name] = value.isoformat()
            else:
                data[name] = value
        return data

    @staticmethod
    def _normalize_row(raw_row: dict) -> dict:
        """Normalize a raw row dict's keys to snake_case for field mapping."""
        normalized: dict[str, Any] = {}
        for key, value in raw_row.items():
            if key is None:
                continue
            norm_key = StagingService._normalize_key(str(key))
            if norm_key:
                normalized[norm_key] = value
        return normalized

    @staticmethod
    def _normalize_key(key: str) -> str:
        """Normalize a header string to snake_case (e.g. 'Account Code' →
        'account_code', 'IFSC' → 'ifsc', 'Opening Balance' → 'opening_balance').
        """
        s = key.strip().lower()
        s = re.sub(r"[\s\-]+", "_", s)
        s = re.sub(r"[^a-z0-9_]", "", s)
        return s

    @staticmethod
    def _to_decimal(value, default: Decimal = ZERO) -> Decimal:
        """Safely convert a value to :class:`Decimal`.

        Handles strings (with thousands separators / currency symbols),
        ints, floats, and existing Decimals. Returns ``default`` on failure
        so storage never crashes on bad data — validation will flag the
        original value via ``raw_data``.
        """
        if value is None:
            return default
        if isinstance(value, Decimal):
            return value
        if isinstance(value, bool):
            return default
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return default
        s = str(value).strip()
        if not s:
            return default
        # Strip currency symbols, spaces, and thousands separators.
        s = re.sub(r"[₹$,\s]", "", s)
        # Handle parentheses as negatives (accounting convention).
        negative = False
        if s.startswith("(") and s.endswith(")"):
            negative = True
            s = s[1:-1].strip()
        try:
            d = Decimal(s)
        except (InvalidOperation, ValueError):
            return default
        return -d if negative else d

    @staticmethod
    def _json_safe_row(raw_row: dict) -> dict:
        """Ensure a raw row dict is JSON-safe for storage in ``raw_data``."""
        safe: dict[str, Any] = {}
        for key, value in raw_row.items():
            if key is None:
                continue
            if isinstance(value, Decimal):
                safe[str(key)] = str(value)
            elif hasattr(value, "isoformat"):
                safe[str(key)] = value.isoformat()
            else:
                safe[str(key)] = value
        return safe

    @staticmethod
    def _log_audit(wizard, action, user, details=None) -> None:
        """Create a :class:`MigrationAuditLog` entry (append-only).

        Wrapped so a logging failure never blocks a legitimate staging
        operation; the error is logged at ERROR level instead.
        """
        try:
            log = MigrationAuditLog(
                wizard=wizard,
                society=wizard.society,
                action=action,
                actor=user,
                details=details or {},
            )
            log.save()
        except Exception:  # noqa: BLE001 — audit must not break the operation.
            logger.exception(
                "Failed to write MigrationAuditLog for wizard %s (action=%s)",
                getattr(wizard, "pk", None),
                action,
            )
